from customtkinter import *
from tkcalendar import DateEntry
from tkinter import messagebox, ttk
from datetime import datetime
from openpyxl import load_workbook
from src.Model.model import StockManagementModel
import os

class StockManagementApp(CTk):
    def __init__(self):
        super().__init__()
        self.model = StockManagementModel()
        self.title("Système de Gestion de Stock")
        self.geometry("800x600")
        self.after(0, lambda: self.state('zoomed'))
        self.current_user_role = None
        self.create_login_interface()
        self.mainloop()

    def create_login_interface(self):
        self.clear_frame()
        self.login_frame = CTkFrame(self)
        self.login_frame.pack(pady=100)
        self.username_entry = self.create_label_entry_button(self.login_frame, label_text="Username:", entry=True, pady=5)
        self.password_entry = self.create_label_entry_button(self.login_frame, label_text="Password:", entry=True, button_text="Login", button_command=self.check_login, show='*', pady=10)

    def create_main_menu(self):
        self.clear_frame()
        self.main_frame = CTkFrame(self)
        self.main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        buttons = [("Add Article", self.add_article),
                   ("Update Article", self.update_article),
                   ("Update Stock", self.update_stock),
                   ("Search Articles", self.search_articles),
                   ("Affichage", self.Affichage),
                   ("Check Low Stock", self.check_low_stock),
                   ("User Management", self.user_management),
                   ("Logout", self.create_login_interface)]

        for text, command in buttons:
            self.create_label_entry_button(self.main_frame, button_text=text, button_command=command, pady=10)

    def create_label_entry_button(self, frame, label_text=None, entry=False, entry_values=None, button_text=None, button_command=None, show=None, pady=10, state="normal"):
        if label_text:
            label = CTkLabel(frame, text=label_text)
            label.pack(pady=pady)
        if entry:
            if entry_values:
                entry_widget = CTkComboBox(frame, values=entry_values, state=state)
            else:
                entry_widget = CTkEntry(frame, show=show, state=state)
            entry_widget.pack(pady=pady)
        if button_text:
            button = CTkButton(frame, text=button_text, command=button_command)
            button.pack(pady=pady)
        if entry:
            return entry_widget
        if button_text:
            return button

    def clear_frame(self):
        for widget in self.winfo_children():
            widget.destroy()

    def check_login(self):
        username = self.username_entry.get()
        password = self.password_entry.get()
        hashed_password = self.model.hash_password(password)
        result = self.model.get_user(username)

        if result and hashed_password == result[0]:
            self.current_user_role = result[1]
            self.create_main_menu()
        else:
            messagebox.showerror("Error", "Invalid username or password")

    def add_article(self):
        self.clear_frame()
        self.add_article_frame = CTkFrame(self)
        self.add_article_frame.pack(fill="both", expand=True, padx=20, pady=20)

        entries = [("Name:", None),
                   ("Category:", self.model.get_categories()),
                   ("Unit Price for Purchase:", None),
                   ("Unit Price for Sale:", None)]
        self.entries = []
        for label_text, values in entries:
            self.entries.append(self.create_label_entry_button(self.add_article_frame, label_text=label_text, entry=True, entry_values=values, pady=5))

        self.create_label_entry_button(self.add_article_frame, button_text="Save", button_command=self.save_article, pady=20)
        self.create_label_entry_button(self.add_article_frame, button_text="Back", button_command=self.create_main_menu, pady=10)

    def save_article(self):
        name, category, price_for_purchase, unit_price = [entry.get() for entry in self.entries]

        if not all([name, category, price_for_purchase, unit_price]):
            messagebox.showerror("Error", "All fields are required!")
            return

        try:
            price_for_purchase = float(price_for_purchase)
            unit_price = float(unit_price)
        except ValueError:
            messagebox.showerror("Error", "Invalid input types!")
            return

        if self.model.add_article(name, category, price_for_purchase, unit_price):
            messagebox.showinfo("Success", "Article added successfully!")
            self.create_main_menu()
        else:
            messagebox.showerror("Error", "Article with this name already exists!")

    def update_article(self):
        articles_names = self.model.get_article_names()
        if not articles_names:
            messagebox.showinfo("The stock is Empty", "The stock is Empty right now!\nTry again after you add some articles.")
            return

        self.clear_frame()
        self.update_times = 0
        self.update_times += 1
        self.update_article_frame = CTkFrame(self)
        self.update_article_frame.pack(fill="both", expand=True, padx=20, pady=20)

        self.name_entry = self.create_label_entry_button(self.update_article_frame, label_text="Name:", entry=True, entry_values=articles_names, pady=5)
        self.create_label_entry_button(self.update_article_frame, button_text="Search", button_command=self.search_article_details, pady=20)
        self.back_button = self.create_label_entry_button(self.update_article_frame, button_text="Back", button_command=self.create_main_menu, pady=10)

    def search_article_details(self):
        name = self.name_entry.get()

        if not name:
            messagebox.showerror("Error", "Article name is required!")
            return

        article = self.model.get_article(name)

        if not article:
            messagebox.showerror("Error", "Article not found!")
            return

        entries = [("Category:", self.model.get_categories(), article[1], "normal"),
                   ("Unit Price for Purchase:", None, article[2], "normal"),
                   ("Unit Price for Sale:", None, article[3], "normal"),
                   ("Quantity:", None, article[4], "readonly")]

        if self.update_times == 1:
            self.back_button.destroy()
            self.update_times += 1

            self.update_entries = []
            for label_text, values, default, state in entries:
                entry = self.create_label_entry_button(self.update_article_frame, label_text=label_text, entry=True, entry_values=values, pady=5)
                if isinstance(entry, CTkEntry):
                    entry.insert(0, default)
                elif isinstance(entry, CTkComboBox):
                    entry.set(default)
                entry.configure(state=state)
                self.update_entries.append(entry)

            self.create_label_entry_button(self.update_article_frame, button_text="Save", button_command=self.save_updated_article, pady=20)
            self.create_label_entry_button(self.update_article_frame, button_text="Delete", button_command=self.delete_article, pady=10)
            self.create_label_entry_button(self.update_article_frame, button_text="Back", button_command=self.create_main_menu, pady=10)
        else:
            for entry, (_, _, default, state) in zip(self.update_entries, entries):
                entry.configure(state="normal")
                if isinstance(entry, CTkEntry):
                    entry.delete(0, 'end')
                    entry.insert(0, default)
                elif isinstance(entry, CTkComboBox):
                    entry.set(default)
                entry.configure(state=state)

    def save_updated_article(self):
        name = self.name_entry.get()
        category, price_for_purchase, unit_price, _ = [entry.get() for entry in self.update_entries]

        if not all([name, category, price_for_purchase, unit_price]):
            messagebox.showerror("Error", "All fields are required!")
            return

        confirmation = messagebox.askyesno("Confirm change the information", f"Do you really want to update the detail for this article '{name}'")
        if not confirmation:
            return

        try:
            price_for_purchase = float(price_for_purchase)
            unit_price = float(unit_price)
        except ValueError:
            messagebox.showerror("Error", "Invalid input types!")
            return

        self.model.update_article(name, category, price_for_purchase, unit_price)
        messagebox.showinfo("Success", "Article updated successfully!")
        self.create_main_menu()

    def delete_article(self):
        if self.current_user_role != 'admin':
            messagebox.showerror("Access Denied", "You do not have permission to delete Articles")
            return

        name = self.name_entry.get()

        if not name:
            messagebox.showerror("Error", "Article name is required!")
            return

        confirmation = messagebox.askyesno("Confirm Delete", f"Do you really want to delete this article: '{name}'")
        if not confirmation:
            return

        self.model.delete_article(name)
        messagebox.showinfo("Success", "Article deleted successfully!")
        self.create_main_menu()

    def update_stock(self):
        articles_names = self.model.get_article_names()
        if not articles_names:
            messagebox.showinfo("The stock is Empty", "The stock is Empty right now!\nTry again after you add some articles.")
            return
    
        self.clear_frame()
        self.update_stock_frame = CTkFrame(self)
        self.update_stock_frame.pack(fill="both", expand=True, padx=20, pady=20)

        self.name_entry = self.create_label_entry_button(self.update_stock_frame, label_text="Name:", entry=True, entry_values=articles_names, pady=5)
        self.reason_entry = self.create_label_entry_button(self.update_stock_frame, label_text="Reason:", entry=True, entry_values=["Sale", "Damage", "Return", "Purchase"], pady=5, state="readonly")
        self.reason_entry.set("Sale")
        self.quantity_entry = self.create_label_entry_button(self.update_stock_frame, label_text="Quantity:", entry=True, pady=5)

        self.create_label_entry_button(self.update_stock_frame, button_text="Update", button_command=self.save_stock_update, pady=20)
        self.create_label_entry_button(self.update_stock_frame, button_text="Back", button_command=self.create_main_menu, pady=10)

    def save_stock_update(self):
        name = self.name_entry.get()
        reason = self.reason_entry.get()
        quantity = self.quantity_entry.get()

        if not all([name, reason, quantity]):
            messagebox.showerror("Error", "All fields are required!")
            return

        try:
            quantity = int(quantity)
        except ValueError:
            messagebox.showerror("Error", "Invalid input types!")
            return

        article = self.model.get_article(name)
        if not article:
            messagebox.showerror("Error", "Article not found!")
            return

        current_quantity = article[4]
        if reason in ["Sale", "Damage"]:
            new_quantity = current_quantity - quantity
            if new_quantity < 0:
                messagebox.showerror("Error", f"Insufficient stock!\nCurrent Quantity: {current_quantity}")
                return
            price = article[3] if reason == "Sale" else article[2]
        elif reason in ["Purchase", "Return"]:
            new_quantity = current_quantity + quantity
            price = article[2] if reason == "Purchase" else article[3]

        confirmation = messagebox.askyesno("Confirm Update", f"Current Quantity: {current_quantity}\nNew Quantity: {new_quantity}")
        if confirmation:
            self.model.update_stock(name, reason, quantity, price)
            messagebox.showinfo("Success", "Stock updated successfully!")
            self.quantity_entry.delete(0, 'end')

    def search_articles(self):
        articles_names = self.model.get_article_names()
        if not articles_names:
            messagebox.showinfo("The stock is Empty", "The stock is Empty right now!\nTry again after you add some articles.")
            return

        self.clear_frame()
        self.search_articles_frame = CTkFrame(self)
        self.search_articles_frame.pack(fill="both", expand=True, padx=20, pady=20)

        entries = [("Category:", lambda: ["ALL"] + self.model.get_categories()),
                   ("Name:", lambda: ["ALL"] + articles_names),
                   ("Status:", lambda: ["ALL", "In stock", "Out of stock"])]

        self.search_entries = []
        self.search_times = 0
        self.search_times += 1
        for label_text, values in entries:
            self.search_entries.append(self.create_label_entry_button(self.search_articles_frame, label_text=label_text, entry=True, entry_values=values(), pady=5))

        self.create_label_entry_button(self.search_articles_frame, button_text="Search", button_command=self.perform_search, pady=20)
        self.create_label_entry_button(self.search_articles_frame, button_text="Back", button_command=self.create_main_menu, pady=10)

    def perform_search(self):
        category, name, status = [entry.get() for entry in self.search_entries]
        articles = self.model.search_articles(category, name, status)

        if self.search_times != 1:
            self.tree.destroy()
        self.tree = ttk.Treeview(self.search_articles_frame, columns=("Name", "Category", "UnitPriceForPurchase", "UnitPriceForSell", "Quantity"), show='headings')
        self.tree.heading("Name", text="Name")
        self.tree.heading("Category", text="Category")
        self.tree.heading("UnitPriceForPurchase", text="Unit Price for Purchase")
        self.tree.heading("UnitPriceForSell", text="Unit Price for Sale")
        self.tree.heading("Quantity", text="Quantity")
        self.tree.pack(pady=10)

        self.search_times += 1

        for article in articles:
            self.tree.insert("", "end", values=article)

    def Affichage(self):
        self.clear_frame()
        self.generate_reports_frame = CTkFrame(self)
        self.generate_reports_frame.pack(fill="both", expand=True, padx=20, pady=20)

        self.report_type_entry = self.create_label_entry_button(self.generate_reports_frame, label_text="Select:", entry=True, entry_values=["Stock", "Statistics"], pady=5, state='readonly')
        self.report_type_entry.set("Stock")
        self.create_label_entry_button(self.generate_reports_frame, button_text="Afficher", button_command=self.Afficher, pady=20)
        self.create_label_entry_button(self.generate_reports_frame, button_text="Back", button_command=self.create_main_menu, pady=10)

    def Afficher(self):
        if self.report_type_entry.get() == "Stock":
            return self.display_report("Stock")

        self.clear_frame()
        self.generate_reports_frame = CTkFrame(self)
        self.generate_reports_frame.pack(fill="both", expand=True, padx=20, pady=20)
        self.start_date_entry = self.create_label_entry_button(self.generate_reports_frame, label_text="Select the start Date:", pady=5)
        self.start_date_entry = DateEntry(self.generate_reports_frame, date_pattern='yyyy-mm-dd')
        self.start_date_entry.pack(pady=5)
        self.end_date_entry = self.create_label_entry_button(self.generate_reports_frame, label_text="Select the End Date:", pady=5)
        self.end_date_entry = DateEntry(self.generate_reports_frame, date_pattern='yyyy-mm-dd')
        self.end_date_entry.pack(pady=5)
        self.create_label_entry_button(self.generate_reports_frame, button_text="Display", button_command=lambda: self.display_report("Statistics"), pady=20)
        self.create_label_entry_button(self.generate_reports_frame, button_text="Back", button_command=self.Affichage, pady=10)

    def display_report(self, report_type, start_date=None, end_date=None):
        if report_type == "Stock":
            columns = ("Name", "Category", "Unit Price for Purchase", "Unit Price for Sale", "Quantity")
            report_data = self.model.get_from_articles()
        elif report_type == "Statistics":
            if not start_date:
                start_date = self.start_date_entry.get_date().strftime('%Y-%m-%d')
                end_date = self.end_date_entry.get_date().strftime('%Y-%m-%d')

            columns = ("OperationName", "ItemName", "Quantity", "UnitPriceForSell_Or_Purchase", "TotalMoney")
            raw_data = self.model.get_from_logs(start_date, end_date)

            report_data = {}
            for row in raw_data:
                operation, name, quantity, price, total = row
                if operation not in report_data:
                    report_data[operation] = []
                report_data[operation].append({
                    'name': name,
                    'quantity': quantity,
                    'price': price,
                    'total': total
                })

        self.clear_frame()
        self.report_frame = CTkFrame(self)
        self.report_frame.pack(fill="both", expand=True, padx=20, pady=20)

        tree = ttk.Treeview(self.report_frame, columns=columns, show='headings')
        for col in columns:
            tree.heading(col, text=col)
        tree.pack(fill="both", expand=True, pady=20)

        if report_type == "Stock":
            for data in report_data:
                tree.insert("", "end", values=data)

            self.export_button = CTkButton(self.report_frame, text="Export to Excel", command=lambda: self.generate_excel_report(report_data, report_type))
        else:
            total_revenue = 0
            for operation, items in report_data.items():
                tree.insert("", "end", values=(operation,))
                for item in items:
                    tree.insert("", "end", values=("", item['name'], item['quantity'], item['price'], item['total']))
                tree.insert("", "end", values=("", "", "", f"{operation} TOTAL", sum(item['total'] for item in items)))
                tree.insert("", "end", values=("",))

                if operation == 'Sale':
                    total_revenue += sum(item['total'] for item in items)
                else:
                    total_revenue -= sum(item['total'] for item in items)
            tree.insert("", "end", values=("", "", "", "TOTAL REVENUE", total_revenue))

            self.Display_detail = CTkButton(self.report_frame, text="Display detail", command=lambda: self.display_detail_statistic(report_data, report_type, start_date, end_date))
            self.Display_detail.pack(pady=20)
            self.export_button = CTkButton(self.report_frame, text="Export to Excel", command=lambda: self.generate_excel_report(report_data, report_type, start_date, end_date))

        self.export_button.pack(pady=20)
        self.create_label_entry_button(self.report_frame, button_text="Back", button_command=self.Affichage, pady=10)

    def display_detail_statistic(self, report_data, report_type, start_date, end_date):
        report_data = self.model.get_logs(start_date, end_date)
        self.clear_frame()
        self.report_frame = CTkFrame(self)
        self.report_frame.pack(fill="both", expand=True, padx=20, pady=20)

        columns = ("ID", "Operation Name", "Name", "Price for sell or buy", "Quantity", "Total Money", "Date")
        tree = ttk.Treeview(self.report_frame, columns=columns, show='headings')
        for col in columns:
            tree.heading(col, text=col)
        tree.column("ID", width=50)
        tree.pack(fill="both", expand=True, pady=20)

        for data in report_data:
            tree.insert("", "end", values=data)

        self.Display_Summary = CTkButton(self.report_frame, text="Display Summary", command=lambda: self.display_report(report_type, start_date, end_date))
        self.export_button = CTkButton(self.report_frame, text="Export to Excel", command=lambda: self.generate_excel_report(report_data, report_type, start_date, end_date))

        self.Display_Summary.pack(pady=20)
        self.export_button.pack(pady=20)
        self.create_label_entry_button(self.report_frame, button_text="Back", button_command=self.Affichage, pady=10)

    def generate_excel_report(self, report_data, report_type, start_date=None, end_date=None):
        os.makedirs('Reports', exist_ok=True)
        if report_type == "Stock":
            template_path = 'Templates/template_inventory_list.xlsx'
            workbook = load_workbook(template_path)
            sheet = workbook.active

            sheet['B3'] = datetime.now().strftime("%Y-%m-%d")

            row_n = 7
            for article in report_data:
                sheet.cell(row=row_n, column=1, value=article[0])  
                sheet.cell(row=row_n, column=2, value=article[1])  
                sheet.cell(row=row_n, column=3, value=article[2])  
                sheet.cell(row=row_n, column=4, value=article[3])  
                sheet.cell(row=row_n, column=5, value=article[4])  
                row_n += 1

            report_file = f"Reports/Stock Report {sheet['B3'].value}.xlsx"
        else:  
            if type(report_data) != list:
                template_path = 'Templates/template_Report_revenue.xlsx'
                workbook = load_workbook(template_path)
                sheet = workbook.active

                sheet.cell(row=2, column=2, value=start_date)
                sheet.cell(row=2, column=3, value=end_date)

                start_row = 4
                current_row = start_row
                total_revenue = 0
                for operation, items in report_data.items():
                    sheet.cell(row=current_row, column=1, value=operation)
                    current_row += 1
                    for item in items:
                        sheet.cell(row=current_row, column=2, value=item['name'])
                        sheet.cell(row=current_row, column=3, value=item['quantity'])
                        sheet.cell(row=current_row, column=4, value=item['price'])
                        sheet.cell(row=current_row, column=5, value=item['total'])
                        current_row += 1
                    sheet.cell(row=current_row, column=4, value=f"{operation} TOTAL")
                    sheet.cell(row=current_row, column=5, value=sum(item['total'] for item in items))

                    if operation == 'Sale':
                        total_revenue += sum(item['total'] for item in items)
                    else:
                        total_revenue -= sum(item['total'] for item in items)

                    current_row += 2

                sheet.cell(row=current_row, column=4, value=f"TOTAL REVENUE")
                sheet.cell(row=current_row, column=5, value=total_revenue)
                report_file = f"Reports/Statistical Summary Report {start_date} - {end_date}.xlsx"

            else:
                template_path = 'Templates/template_statistique_detail.xlsx'
                workbook = load_workbook(template_path)
                sheet = workbook.active

                sheet.cell(row=2, column=3, value=start_date)
                sheet.cell(row=2, column=4, value=end_date)

                current_row = 4
                for item in report_data:
                    sheet.cell(row=current_row, column=1, value=item[0])
                    sheet.cell(row=current_row, column=2, value=item[1])
                    sheet.cell(row=current_row, column=3, value=item[2])
                    sheet.cell(row=current_row, column=4, value=item[3])
                    sheet.cell(row=current_row, column=5, value=item[4])
                    sheet.cell(row=current_row, column=6, value=item[5])
                    sheet.cell(row=current_row, column=7, value=item[6])
                    current_row += 1

                report_file = f"Reports/Statistique Detailed Report {start_date} - {end_date}.xlsx"

        workbook.save(report_file)
        report_file = report_file.split("/")[-1]
        messagebox.showinfo("Success", f"Report generated: {report_file}")

    def check_low_stock(self):
        self.clear_frame()
        self.low_stock_frame = CTkFrame(self)
        self.low_stock_frame.pack(fill="both", expand=True, padx=20, pady=20)

        self.threshold_entry = self.create_label_entry_button(self.low_stock_frame, label_text="Stock Threshold:", entry=True, pady=5)
        self.threshold_entry.insert(0, '20')

        self.create_label_entry_button(self.low_stock_frame, button_text="Check", button_command=self.show_low_stock, pady=20)
        self.create_label_entry_button(self.low_stock_frame, button_text="Back", button_command=self.create_main_menu, pady=10)

    def show_low_stock(self):
        threshold = self.threshold_entry.get()

        try:
            threshold = int(threshold)
        except ValueError:
            messagebox.showerror("Error", "Invalid input type!")
            return

        low_stock_materials = self.model.check_low_stock(threshold)

        if not low_stock_materials:
            messagebox.showinfo("No Results", f"No materials found with quantity under {threshold}!")
            return

        self.clear_frame()
        self.check_low_stock_frame = CTkFrame(self)
        self.check_low_stock_frame.pack(fill="both", expand=True, padx=20, pady=20)

        tree = ttk.Treeview(self.check_low_stock_frame, columns=("Name", "Category", "Unit Price for Purchase", "Unit Price for Sale", "Quantity"), show='headings')
        tree.heading("Name", text="Name")
        tree.heading("Category", text="Category")
        tree.heading("Unit Price for Purchase", text="Unit Price for Purchase")
        tree.heading("Unit Price for Sale", text="Unit Price for Sale")
        tree.heading("Quantity", text="Quantity")
        tree.pack(pady=10)

        for article in low_stock_materials:
            tree.insert("", "end", values=(article[0], article[1], article[2], article[3], article[4]))

        self.create_label_entry_button(self.check_low_stock_frame, button_text="Back", button_command=self.check_low_stock, pady=10)

    def user_management(self):
        if self.current_user_role != 'admin':
            messagebox.showerror("Access Denied", "You do not have permission to access user management")
            return

        self.clear_frame()
        self.user_management_frame = CTkFrame(self)
        self.user_management_frame.pack(pady=20)

        self.new_username_entry = self.create_label_entry_button(self.user_management_frame, label_text="Username", entry=True, pady=5)
        self.new_password_entry = self.create_label_entry_button(self.user_management_frame, label_text="Password", entry=True, pady=5)
        self.role_entry = self.create_label_entry_button(self.user_management_frame, label_text="Role", entry=True, entry_values=["admin", "user"], pady=5, state="readonly")
        self.role_entry.set("user")

        self.create_label_entry_button(self.user_management_frame, button_text="Create User", button_command=self.create_user, pady=20)
        self.create_label_entry_button(self.user_management_frame, button_text="Delete User", button_command=self.delete_user, pady=10)
        self.create_label_entry_button(self.user_management_frame, button_text="Back", button_command=self.create_main_menu, pady=10)

    def create_user(self):
        username = self.new_username_entry.get()
        password = self.new_password_entry.get()
        role = self.role_entry.get()

        if not username or not password:
            messagebox.showerror("Error", "All fields are required!")
            return

        if self.model.create_user(username, password, role):
            messagebox.showinfo("Success", "User created successfully")
            self.new_username_entry.delete(0, 'end')
            self.new_password_entry.delete(0, 'end')
        else:
            messagebox.showerror("Error", "User with this username already exists!")

    def delete_user(self):
        username = self.new_username_entry.get()

        if not username:
            messagebox.showerror("Error", "username is required!")
            return

        confirmation = messagebox.askyesno("Confirm Delete", f"Do you really want to delete this user: '{username}'")
        if not confirmation:
            return

        user_exists = self.model.get_user(username)
        if not user_exists:
            messagebox.showerror("Error", "User with this username Doesn't exist!")
            return

        self.model.delete_user(username)
        messagebox.showinfo("Success", "User deleted successfully!")
        self.new_username_entry.delete(0, 'end')
